import logging
import re
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Prefetch
from django.shortcuts import render, redirect, get_object_or_404

from .models import Cliente, Producto, Remision, Venta, DetalleVenta
from .forms import RemisionForm, VentaForm, DetalleVentaFormSet

logger = logging.getLogger(__name__)


# -----------------------------
# HOME
# -----------------------------
def home(request):
    return render(request, "sistema/home.html")


# -----------------------------
# HELPERS
# -----------------------------
def safe_decimal(val) -> Decimal:
    """Convierte valores de Excel a Decimal sin reventar."""
    if val is None:
        return Decimal("0")
    s = str(val).strip()
    if s in ("", "-", "na", "NA", "N/A", "None", "nan"):
        return Decimal("0")
    s = s.replace(",", "")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def safe_int(val, default=0) -> int:
    """Convierte valores a int tolerando Excel sucio (#, 12.0, etc)."""
    if val is None:
        return default
    s = str(val).strip()

    if s in ("", "-", "#", "N/A", "NA", "nan", "None"):
        return default

    # Excel a veces manda 12.0
    try:
        return int(float(s))
    except Exception:
        pass

    # último intento: quedarte con dígitos
    digits = "".join(ch for ch in s if ch.isdigit())
    return int(digits) if digits else default


def safe_str(val) -> str:
    return str(val).strip() if val is not None else ""


# -----------------------------
# IMPORTAR PRODUCTOS
# -----------------------------
def importar_productos(request):
    if request.method == "POST":
        try:
            archivo = request.FILES.get("excel_file")
            if not archivo:
                messages.error(request, "No se recibió ningún archivo. Revisa que el input se llame excel_file.")
                return redirect("sistema:importar_productos")

            wb = load_workbook(archivo, data_only=True)
            ws = wb.active

            primera_fila = True
            segunda_fila = True

            with transaction.atomic():
                for row in ws.iter_rows(values_only=True):
                    if primera_fila:
                        primera_fila = False
                        continue
                    if segunda_fila:
                        segunda_fila = False
                        continue

                    if not row or all(col is None for col in row):
                        continue

                    # Evita IndexError si faltan columnas
                    if len(row) < 7:
                        continue

                    codigo = row[1]
                    descripcion = row[2]
                    compra_cjs = safe_decimal(row[3])
                    compra_pzs = safe_decimal(row[4])
                    venta_cjs = safe_decimal(row[5])
                    venta_pzs = safe_decimal(row[6])

                    if not codigo:
                        continue

                    Producto.objects.update_or_create(
                        codigo=safe_str(codigo),
                        defaults={
                            "descripcion": safe_str(descripcion),
                            "compra_cjs": compra_cjs,
                            "compra_pzs": compra_pzs,
                            "venta_cjs": venta_cjs,
                            "venta_pzs": venta_pzs,
                        },
                    )

            messages.success(request, "Productos importados correctamente.")
            return redirect("sistema:lista_productos")

        except Exception:
            logger.exception("ERROR EN IMPORTAR PRODUCTOS")
            messages.error(request, "Ocurrió un error al importar. Revisa los logs en Render.")
            return redirect("sistema:importar_productos")

    return render(request, "sistema/importar_productos.html")


# -----------------------------
# IMPORTAR CLIENTES (ROBUSTO)
# -----------------------------
def importar_clientes(request):
    if request.method == "POST":
        archivo = request.FILES.get("excel_file")
        if not archivo:
            messages.error(
                request,
                "No se recibió el archivo. Revisa que el formulario tenga enctype y el input se llame excel_file.",
            )
            return redirect("sistema:importar_clientes")

        try:
            wb = load_workbook(archivo, data_only=True)
            ws = wb.active

            primera_fila = True
            segunda_fila = True
            creados = 0
            actualizados = 0

            with transaction.atomic():
                for row in ws.iter_rows(values_only=True):
                    if primera_fila:
                        primera_fila = False
                        continue
                    if segunda_fila:
                        segunda_fila = False
                        continue

                    if not row or all(col is None for col in row):
                        continue

                    # Asegura columnas 0..6
                    row = list(row) + [None] * (7 - len(row))

                    numero_raw = row[0]
                    proveedor = safe_str(row[1])
                    comercio = safe_str(row[2])
                    contacto = safe_str(row[3])
                    direccion = safe_str(row[4])
                    telefono = safe_str(row[5])
                    referencia = safe_str(row[6])

                    if not proveedor:
                        continue

                    numero = safe_int(numero_raw, default=0)

                    _, created = Cliente.objects.update_or_create(
                        proveedor=proveedor,
                        defaults={
                            "numero": numero,
                            "comercio": comercio,
                            "contacto": contacto,
                            "direccion": direccion,
                            "telefono": telefono,
                            "referencia": referencia,
                        },
                    )

                    if created:
                        creados += 1
                    else:
                        actualizados += 1

            messages.success(
                request,
                f"Clientes importados correctamente. Nuevos: {creados} | Actualizados: {actualizados}",
            )
            return redirect("sistema:lista_clientes")

        except Exception as e:
            logger.exception("ERROR EN IMPORTAR CLIENTES")
            messages.error(request, f"Error al importar clientes: {e}")
            return redirect("sistema:importar_clientes")

    return render(request, "sistema/importar_clientes.html")


# -----------------------------
# LISTADOS
# -----------------------------
def lista_productos(request):
    productos = Producto.objects.all().order_by("codigo")
    return render(request, "sistema/lista_productos.html", {"productos": productos})


def lista_clientes(request):
    clientes = Cliente.objects.all().order_by("comercio")
    return render(request, "sistema/lista_clientes.html", {"clientes": clientes})


# -----------------------------
# BÚSQUEDA GLOBAL
# -----------------------------
def busqueda_global(request):
    q = request.GET.get("q", "").strip()

    productos = Producto.objects.none()
    clientes = Cliente.objects.none()

    if q:
        productos = Producto.objects.filter(Q(codigo__icontains=q) | Q(descripcion__icontains=q))
        clientes = Cliente.objects.filter(
            Q(proveedor__icontains=q)
            | Q(comercio__icontains=q)
            | Q(contacto__icontains=q)
            | Q(direccion__icontains=q)
        )

    return render(request, "sistema/busqueda.html", {"q": q, "productos": productos, "clientes": clientes})


# -----------------------------
# REMISIONES
# -----------------------------
def remision_list(request):
    remisiones = Remision.objects.select_related("cliente").order_by("-fecha", "-id")
    return render(request, "sistema/remisiones_list.html", {"remisiones": remisiones})


def remision_create(request):
    if request.method == "POST":
        form = RemisionForm(request.POST, request.FILES)
        if form.is_valid():
            remision = form.save()
            return redirect("sistema:remision_detail", pk=remision.pk)
    else:
        form = RemisionForm()

    return render(request, "sistema/remision_form.html", {"form": form})


def remision_detail(request, pk):
    remision = get_object_or_404(Remision.objects.select_related("cliente"), pk=pk)
    return render(request, "sistema/remision_detail.html", {"remision": remision})


# -----------------------------
# VENTAS CRUD
# -----------------------------
def venta_list(request):
    ventas = Venta.objects.select_related("remision", "remision__cliente").order_by("-fecha", "-id")
    return render(request, "sistema/ventas_list.html", {"ventas": ventas})


@transaction.atomic
def venta_create_from_remision(request, remision_id):
    remision = get_object_or_404(Remision.objects.select_related("cliente"), pk=remision_id)

    if hasattr(remision, "venta"):
        return redirect("sistema:venta_edit", pk=remision.venta.pk)

    venta = Venta.objects.create(
        remision=remision,
        fecha=remision.fecha,
        descuento=Decimal("0.00"),
        iva=Decimal("0.00"),
        subtotal=Decimal("0.00"),
        total=Decimal("0.00"),
    )
    return redirect("sistema:venta_edit", pk=venta.pk)


def venta_detail(request, pk):
    venta = get_object_or_404(Venta.objects.select_related("remision", "remision__cliente"), pk=pk)
    detalles = venta.detalles.select_related("producto").all()
    return render(request, "sistema/venta_detail.html", {"venta": venta, "detalles": detalles})


@transaction.atomic
def venta_edit(request, pk):
    venta = get_object_or_404(Venta.objects.select_related("remision", "remision__cliente"), pk=pk)

    if request.method == "POST":
        form = VentaForm(request.POST, instance=venta)
        formset = DetalleVentaFormSet(request.POST, instance=venta)

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            venta.recalcular_totales(commit=True)
            messages.success(request, "Venta actualizada correctamente.")
            return redirect("sistema:venta_detail", pk=venta.pk)
    else:
        form = VentaForm(instance=venta)
        formset = DetalleVentaFormSet(instance=venta)

    return render(request, "sistema/venta_edit.html", {"venta": venta, "form": form, "formset": formset})


# -----------------------------
# IMPORTAR REMISIONES DESDE EXCEL
# -----------------------------
def importar_remisiones_excel(request):
    if request.method == "POST":
        archivo = request.FILES.get("excel_file")
        if not archivo:
            return render(request, "sistema/importar_remisiones.html", {"error": "No se subió archivo."})

        wb = load_workbook(archivo, data_only=True)

        sheet_name = "REL REM ENTREG1"
        if sheet_name not in wb.sheetnames:
            return render(
                request,
                "sistema/importar_remisiones.html",
                {"error": f"No encontré la hoja '{sheet_name}'. Hojas: {wb.sheetnames}"},
            )

        ws = wb[sheet_name]

        header_row = 5
        date_map = {}

        mon_map = {
            "Ene": 1, "Feb": 2, "Mar": 3, "Abr": 4, "May": 5, "Jun": 6,
            "Jul": 7, "Ago": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dic": 12,
            "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
            "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
        }

        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if isinstance(val, str):
                m = re.search(r"(\d{2})/([A-Za-z]{3})/(\d{2})", val)
                if m:
                    dd, mon, yy = m.groups()
                    month = mon_map.get(mon)
                    if month:
                        year = 2000 + int(yy)
                        date_map[col] = date(year, month, int(dd))

        if not date_map:
            return render(request, "sistema/importar_remisiones.html", {"error": "No pude detectar columnas con fechas en la fila 5."})

        creadas = 0
        ya_existian = 0
        ventas_creadas = 0

        for row in range(6, ws.max_row + 1):
            clave_cte = ws.cell(row=row, column=3).value
            comercio = ws.cell(row=row, column=4).value
            contacto = ws.cell(row=row, column=5).value

            if not clave_cte:
                continue

            clave_cte = safe_str(clave_cte)
            comercio = safe_str(comercio)
            contacto = safe_str(contacto)

            cliente, _ = Cliente.objects.get_or_create(
                proveedor=clave_cte,
                defaults={
                    "numero": 0,
                    "comercio": comercio,
                    "contacto": contacto,
                    "direccion": "",
                    "telefono": "",
                    "referencia": "",
                },
            )

            for col, fecha in date_map.items():
                cell_val = ws.cell(row=row, column=col).value
                if not cell_val:
                    continue

                if isinstance(cell_val, str) and "remision" in cell_val.lower():
                    folio = cell_val.replace("Remision", "").replace("remision", "").strip()
                    if not folio:
                        continue

                    remision, created = Remision.objects.get_or_create(
                        cliente=cliente,
                        folio=folio,
                        defaults={"fecha": fecha, "observaciones": ""},
                    )

                    if created:
                        creadas += 1
                        if not hasattr(remision, "venta"):
                            Venta.objects.create(
                                remision=remision,
                                fecha=remision.fecha,
                                subtotal=Decimal("0.00"),
                                total=Decimal("0.00"),
                                descuento=Decimal("0.00"),
                                iva=Decimal("0.00"),
                            )
                            ventas_creadas += 1
                    else:
                        ya_existian += 1

        return render(
            request,
            "sistema/importar_remisiones.html",
            {"ok": True, "creadas": creadas, "ya_existian": ya_existian, "ventas_creadas": ventas_creadas},
        )

    return render(request, "sistema/importar_remisiones.html")


# -----------------------------
# VENTAS (FILTRO POR CLIENTE Y PRODUCTO)
# -----------------------------
def ventas_lista(request):
    cliente_id = request.GET.get("cliente")
    producto_id = request.GET.get("producto")

    qs = (
        Venta.objects.select_related("remision", "remision__cliente")
        .prefetch_related(
            Prefetch(
                "detalles",
                queryset=DetalleVenta.objects.select_related("producto").order_by("id"),
            )
        )
        .order_by("-fecha", "-id")
    )

    if cliente_id:
        qs = qs.filter(remision__cliente_id=cliente_id)

    if producto_id:
        qs = qs.filter(detalles__producto_id=producto_id).distinct()

    context = {
        "ventas": qs[:500],
        "clientes": Cliente.objects.order_by("comercio"),
        "productos": Producto.objects.order_by("codigo"),
        "cliente_sel": cliente_id or "",
        "producto_sel": producto_id or "",
    }
    return render(request, "sistema/ventas_lista.html", context)
